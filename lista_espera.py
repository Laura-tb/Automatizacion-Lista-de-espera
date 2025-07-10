from playwright.sync_api import sync_playwright
import datetime
import openpyxl
import re
import smtplib
from email.message import EmailMessage
import json

# Cargar datos desde config.json
with open("config.example.json", encoding="utf-8") as f:
    config = json.load(f)

# CONFIGURACIÓN: Tus datos personales
FECHA_NACIMIENTO = config["fecha_nacimiento"]  # Formato DD/MM/AAAA
CODIGO = config["codigo"]  # 4 bloques de 5 dígitos

# NOMBRE DEL ARCHIVO EXCEL
EXCEL_FILENAME = "lista_espera.xlsx"

# CONFIGURACIÓN DEL EMAIL
TU_EMAIL = config["email"]  
TU_CONTRASEÑA = config["contrasena"]  
DESTINATARIO = config["destinatario"]  

def extraer_numero(texto):
    match = re.search(r"es de (\d+)", texto)
    return int(match.group(1)) if match else None

def enviar_email(anterior, actual):
    msg = EmailMessage()
    msg['Subject'] = '🔔 Cambio en lista de espera quirúrgica'
    msg['From'] = TU_EMAIL
    msg['To'] = DESTINATARIO
    msg.set_content(f'⚠️ El número de personas en espera ha cambiado de {anterior} a {actual}.')

    try:
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
            smtp.login(TU_EMAIL, TU_CONTRASEÑA)
            smtp.send_message(msg)
        print("Correo enviado con éxito.")
    except Exception as e:
        print(f"Error al enviar el correo: {e}")

def automatizar_lista_espera():
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()

        # 1. Ir al login
        page.goto("https://www.sanidadmadrid.org:444/lespera/plogin.jsp")

        # 2. Rellenar el formulario
        page.fill('input[name="v_fNacimiento"]', FECHA_NACIMIENTO)
        page.fill('input[name="v_codigo1"]', CODIGO[0])
        page.fill('input[name="v_codigo2"]', CODIGO[1])
        page.fill('input[name="v_codigo3"]', CODIGO[2])
        page.fill('input[name="v_codigo4"]', CODIGO[3])

        # 3. Acceder
        page.wait_for_selector('button.button-red', timeout=60000)
        page.click('button.button-red')
        page.wait_for_load_state("networkidle")

        # 4. Esperar y extraer el texto
        texto = page.inner_text("body")
        numero = extraer_numero(texto)

        if numero is None:
            print("No se encontró el número de personas en espera.")
        else:
            print(f"Número de personas en espera con mayor demora: {numero}")

            # 5. Guardar con fecha y hora
            fecha_hora = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            try:
                wb = openpyxl.load_workbook(EXCEL_FILENAME)
            except FileNotFoundError:
                wb = openpyxl.Workbook()
            ws = wb.active
            ws.append([fecha_hora, numero])
            wb.save(EXCEL_FILENAME)
            print(f"Guardado en {EXCEL_FILENAME}")

            # 6. Comparar con el anterior y enviar email si cambia
            filas = list(ws.iter_rows(values_only=True))
            if len(filas) >= 2:
                anterior = filas[-2][1]
                actual = filas[-1][1]
                if anterior != actual:
                    enviar_email(anterior, actual)

        browser.close()

# Ejecutar
automatizar_lista_espera()
